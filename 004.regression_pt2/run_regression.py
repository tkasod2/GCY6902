# -*- coding: utf-8 -*-
# run_regression.py
import torch
import os
import re
import gc
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import load_workbook
import pickle
import matplotlib.pyplot as plt
import seaborn as sns

from torch.utils.data import DataLoader

# 기존 파일에서 클래스 및 함수 불러오기
from tft_model import TFTConfig, TemporalFusionTransformer
from train_tft import SeqDataset, collate_fn, train, train_lstm
from sklearn.metrics import mean_squared_error, mean_absolute_error
from sklearn.metrics import accuracy_score, f1_score
from sklearn.metrics import roc_auc_score, confusion_matrix, roc_curve, auc
from sklearn.preprocessing import label_binarize
from prepare_data import *
import random

# 모델날짜
today_ = datetime.now().strftime("%y%m%d")

# seed 전체 고정
def seed_worker(worker_id):
    worker_seed = torch.initial_seed() % 2**32

def static_seed(seed = 42):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.use_deterministic_algorithms(True)

def fit_and_out(df,
                target,
                outdir,
                config_dict: dict,
                var_list: list,
                grp_cd="Symbol",
                lr = 5e-4, 
                epochs=2,
                patience=1,
                trk_start_dt='20240101',  # 시작 일자
                seq_length=60, # 60개월
                dt_cut='20251231',         # train/valid 분할 기준 일자
                threshold=0.05, # 변동 방향성 판단 기준 (예: 5% 이상 상승/하락)
                tgt_gap=3,
                seed=42,
                train_yn = 1):
    """
    * 학습용 param
    df : 모델 학습용 데이터프레임
    target : 결과변수 설정
    config_dict : 모델 configure
    var_list : past_vars 목록
    grp_cd : 집계 단위 -> 산업군 78 개 (260424)
    epochs : epoch
    patience : 몇회참고 early stopping 할지

    * output용 param
    outdir : 아웃풋 경로
    threshold : 3-classification task처럼 바꿔서 정확도 측정을 위함

    * 전처리용 param
    trk_start_dt : 분석시작연월정보
    seq_length : 1 seq 길이
    dt_cut : validation 마지막 연월정보
    tgt_gap : 몇 개월후를 추정할지에 대한 시차

    * 기타
    seed : 재현성 확보 목적 seed 고정
    """
    static_seed(seed)

    x_past, x_known, x_static, y, scaler, le, target_scalers = prepare_data(
        df,
        target=target,
        var_list=var_list,
        trk_start_dt=trk_start_dt,
        seq_length=seq_length,
        dt_cut=dt_cut,
        tgt_gap=tgt_gap
    )

    # labelencoder, StandardScaler 저장
    with open(outdir+'label_encoder.pkl','wb') as f:
        pickle.dump(le, f, pickle.HIGHEST_PROTOCOL)
    with open(outdir+'scaler.pkl','wb') as f:
        pickle.dump(scaler, f, pickle.HIGHEST_PROTOCOL)
    with open(outdir+'target_scalers.pkl','wb') as f:
        pickle.dump(target_scalers, f, pickle.HIGHEST_PROTOCOL)

    df_scaler = pd.DataFrame({
        "VAR": scaler.feature_names_in_,
        "mean": scaler.mean_,
        "std": scaler.scale_
    })
    # mean_scale = df_scaler.loc[df_scaler['VAR'] == target, 'mean'].item()
    # std_scale = df_scaler.loc[df_scaler['VAR'] == target, 'std'].item()

    # ynum = var_list.index(target)
    # y_t_minus_1_raw = x_past[2][:, -1, ynum] # 테스트셋의 t-1 시점 원본값
    # y_t_minus_1_raw = y_t_minus_1_raw * std_scale + mean_scale

    ynum = var_list.index(target)
    y_t_minus_1_scaled = x_past[2][:, -1, ynum] # 테스트셋의 t-1 시점 스케일된 값

    # 재현 가능성 고려하여, Seed 고정
    g = torch.Generator()
    g.manual_seed(seed)

    # 데이터 Sqeunce 처리 및 Dataloader 세팅
    train_ds = SeqDataset(x_past[0], y[0], x_known[0], x_static[0])
    valid_ds = SeqDataset(x_past[1], y[1], x_known[1], x_static[1])
    test_ds = SeqDataset(x_past[2], y[2], x_known[2], x_static[2])
    batch_size = 128
    trn_loader = DataLoader(train_ds, batch_size=batch_size, shuffle=True
                            , collate_fn=collate_fn, worker_init_fn = seed_worker, generator=g, num_workers=2,pin_memory=True)
    val_loader = DataLoader(valid_ds, batch_size=batch_size, shuffle=False, collate_fn=collate_fn, num_workers=2,pin_memory=True)
    ts_loader = DataLoader(test_ds, batch_size=batch_size, shuffle=False, collate_fn=collate_fn, num_workers=2,pin_memory=True)

    def fit_model(config_dict,epochs=2,patience=1):
        cfg = TFTConfig(
            **config_dict
        )

        print("=======1. TFT 모델 학습중=======")
        model, df_epoch_tft = train(
            cfg=cfg,
            train_loader=trn_loader,
            valid_loader=val_loader,
            test_loader=ts_loader,
            epochs=epochs,
            lr=lr,
            patience=patience,
            save_path=outdir + f'{target}_tft_test_{today_}.pt'
        )

        print("=======2. LSTM 모델 학습중=======")
        model_lstm, df_epoch_lstm = train_lstm(
            input_dim=cfg.past_vars,
            hidden_dim=cfg.lstm_hidden,
            train_loader=trn_loader,
            valid_loader=val_loader,
            test_loader=ts_loader,
            epochs=epochs,
            lr=lr,
            patience=patience,
            save_path=outdir + f'{target}_lstm_test_{today_}.pt'
        )
        print(config_dict)

        df_epoch_tft['gb'] = 'TFT'
        df_epoch_lstm['gb'] = 'LSTM'

        df_epoch = pd.concat([df_epoch_tft, df_epoch_lstm], axis=0, ignore_index=True)
        df_config = pd.DataFrame.from_dict({'keys': config_dict.keys(),'values': config_dict.values()})
        df_epoch.to_csv(outdir+'df_epoch.csv.gz',compression='gzip', index=False)
        df_config.to_csv(outdir+'df_config.csv.gz',compression='gzip', index=False)
        gc.collect()

        return model, model_lstm, df_epoch, df_config

    # 학습을 할지 혹은 기존 모델로 추론만 할지 결정
    if train_yn==1: # 학습만 수행시
        model, model_lstm, df_epoch, df_config = fit_model(config_dict,epochs=epochs,patience=patience)
        excel_path = outdir + f'{target}_final_report_{today_}.xlsx'
    else: # 추론만 수행시
        date_pattern_ = re.search(r'\d{6}', outdir)
        dt_learn = date_pattern_.group()

        model = outdir + target +f"_tft_test_{dt_learn}.pt"
        model_lstm = outdir + target +f"_lstm_test_{dt_learn}.pt"
        df_epoch = pd.read_csv(outdir+'df_epoch.csv.gz')
        df_config = pd.DataFrame.from_dict({'keys': config_dict.keys(),'values': config_dict.values()})
        model, model_lstm, df_epoch, df_config = fit_model(config_dict,epochs=epochs,patience=patience)
        excel_path = outdir + f'{target}_infer_report_{dt_learn}.xlsx'

    # 3. 예측 수행
    print("===========================output building 중===========================")
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = model.to(device)
    model_lstm = model_lstm.to(device)
    
    model.eval()
    model_lstm.eval()
    with torch.no_grad():
        # 기존 저장 리스트 + 분위수 3개용 리스트 분리 선언
        all_actuals, all_lstm_preds, all_weights, all_attention_score = [], [], [], []
        all_tft_q10, all_tft_q50, all_tft_q90 = [], [], [] 

        for x_p, x_k, x_s, p, y_reg, y_cls in ts_loader:
            x_p = x_p.to(device)
            if x_k is not None: x_k = x_k.to(device)
            if x_s is not None: x_s = x_s.to(device)
            if p is not None: p = p.to(device)

            # tft 결과는 튜플: (y_hat_reg, y_hat_cls)
            y_tft_tuple, aux = model(x_p, x_known=x_k, x_static=x_s, pos=p)
            y_tft_reg = y_tft_tuple[0] # (B, 3) 형태의 분위수 예측값
            
            # lstm
            y_lstm, _ = model_lstm(x_p)

            # --- 분위수 분리 저장 (0.1, 0.5, 0.9) ---
            all_tft_q10.append(y_tft_reg[:, 0].cpu().numpy())
            all_tft_q50.append(y_tft_reg[:, 1].cpu().numpy()) # 이게 기존의 all_tft_preds 역할 (중앙값)
            all_tft_q90.append(y_tft_reg[:, 2].cpu().numpy())

            # --- 기존 로직 완벽 유지 ---
            all_actuals.append(y_reg.numpy())
            all_lstm_preds.append(y_lstm.cpu().numpy())
            all_weights.append(aux['w_past'].cpu().numpy())             # 변수 중요도
            all_attention_score.append(aux['attn_w'].cpu().numpy())     # 어텐션 스코어

    # 결과 data frame 구성
    res_df = pd.DataFrame({
        'TIME_IDX': [x+tgt_gap+1 for x in test_ds.x_known[:, -1, :]],
        grp_cd: le.inverse_transform(test_ds.x_static.reshape(-1,)),
        'Actual': np.concatenate(all_actuals).flatten(),
        'TFT_Q10': np.concatenate(all_tft_q10).flatten(),
        'TFT_Prediction': np.concatenate(all_tft_q50).flatten(), # 0.5 분위수를 대표값으로 사용
        'TFT_Q90': np.concatenate(all_tft_q90).flatten(),
        'LSTM_Prediction': np.concatenate(all_lstm_preds).flatten(),
        'Actual_t_minus_1_scaled': y_t_minus_1_scaled # (이전 대화에서 그룹별 스케일링을 위해 저장해둔 t-1)
    })
    # 결과 data frame

    # res_df.to_csv(outdir + f"{target}_comparison_results.csv", index=False)
    print("===========RAW기준으로 데이터 복원===========")
    res_df_raw = res_df.copy()

    # (이전 대화에서 수정한) 각 행의 산업코드(grp_cd)에 맞는 mean, std 매핑
    res_df_raw['mean_scale'] = res_df_raw[grp_cd].map(lambda x: target_scalers[x].mean_[0])
    res_df_raw['std_scale'] = res_df_raw[grp_cd].map(lambda x: target_scalers[x].scale_[0])

    # 새로 추가된 Q10, Q90을 포함하여 모두 그룹별 스케일 역변환 수행
    res_df_raw['Actual'] = res_df_raw['Actual'] * res_df_raw['std_scale'] + res_df_raw['mean_scale']
    res_df_raw['TFT_Q10'] = res_df_raw['TFT_Q10'] * res_df_raw['std_scale'] + res_df_raw['mean_scale']
    res_df_raw['TFT_Prediction'] = res_df_raw['TFT_Prediction'] * res_df_raw['std_scale'] + res_df_raw['mean_scale']
    res_df_raw['TFT_Q90'] = res_df_raw['TFT_Q90'] * res_df_raw['std_scale'] + res_df_raw['mean_scale']
    res_df_raw['LSTM_Prediction'] = res_df_raw['LSTM_Prediction'] * res_df_raw['std_scale'] + res_df_raw['mean_scale']
    res_df_raw['Actual_t_minus_1'] = res_df_raw['Actual_t_minus_1_scaled'] * res_df_raw['std_scale'] + res_df_raw['mean_scale']
    # 계산에만 쓴 임시 컬럼은 삭제 (선택 사항)
    res_df_raw.drop(columns=['mean_scale', 'std_scale', 'Actual_t_minus_1_scaled'], inplace=True)

    threshold = threshold # +-5% 기준
    def classify_movement(current, base, thr):
        diff = (current - base) / (base + 1e-9) # 분모 0 방지
        if diff > thr: return "상승"
        elif diff < -thr: return "하락"
        else: return "유지"

    # 실제 변동(Actual t vs Actual t-1)
    res_df_raw['Actual_Move'] = res_df_raw.apply(
        lambda x: classify_movement(x['Actual'], x['Actual_t_minus_1'], threshold), axis=1)

    # TFT 예측 변동(TFT Pred t vs Actual t-1)
    res_df_raw['TFT_Move'] = res_df_raw.apply(
        lambda x: classify_movement(x['TFT_Prediction'], x['Actual_t_minus_1'], threshold), axis=1)

    # LSTM 예측 변동(LSTM Pred t vs Actual t-1)
    res_df_raw['LSTM_Move'] = res_df_raw.apply(
        lambda x: classify_movement(x['LSTM_Prediction'], x['Actual_t_minus_1'], threshold), axis=1)

    # Confusion Matrix 계산
    move_labels = ["하락", "유지", "상승"]
    cm_tft = confusion_matrix(res_df_raw['Actual_Move'], res_df_raw['TFT_Move'], labels=move_labels)
    cm_lstm = confusion_matrix(res_df_raw['Actual_Move'], res_df_raw['LSTM_Move'], labels=move_labels)

    # 정확도 계산
    acc_tft = accuracy_score(res_df_raw['Actual_Move'], res_df_raw['TFT_Move'])
    acc_lstm = accuracy_score(res_df_raw['Actual_Move'], res_df_raw['LSTM_Move'])

    print(f"TFT Directional Accuracy: {acc_tft:.4f}")
    print(f"LSTM Directional Accuracy: {acc_lstm:.4f}")

    actual = res_df['Actual']
    tft_pred = res_df['TFT_Prediction']
    lstm_pred = res_df['LSTM_Prediction']

    tft_rmse = np.sqrt(mean_squared_error(actual, tft_pred))
    tft_mae = mean_absolute_error(actual, tft_pred)

    lstm_rmse = np.sqrt(mean_squared_error(actual, lstm_pred))
    lstm_mae = mean_absolute_error(actual, lstm_pred)

    print(tft_rmse, tft_mae, lstm_rmse, lstm_mae)

    tft_raw_rmse = np.sqrt(mean_squared_error(res_df_raw['Actual'], res_df_raw['TFT_Prediction']))
    tft_raw_mae = mean_absolute_error(res_df_raw['Actual'], res_df_raw['TFT_Prediction'])
    tft_acc = accuracy_score(res_df_raw['Actual_Move'], res_df_raw['TFT_Move'])

    lstm_raw_rmse = np.sqrt(mean_squared_error(res_df_raw['Actual'], res_df_raw['LSTM_Prediction']))
    lstm_raw_mae = mean_absolute_error(res_df_raw['Actual'], res_df_raw['LSTM_Prediction'])
    lstm_acc = accuracy_score(res_df_raw['Actual_Move'], res_df_raw['LSTM_Move'])

    def make_cm_df(cm, model_name):
        df_cm = pd.DataFrame(cm, index=[f'Actual_{x}' for x in move_labels], 
                             columns=[f'Pred_{x}' for x in move_labels])
        df_cm.insert(0, 'Model', model_name)
        return df_cm
    df_cm_tft = make_cm_df(cm_tft, 'TFT')
    df_cm_lstm = make_cm_df(cm_lstm, 'LSTM')
    
    # 두 모델의 CM을 위아래로 이어붙임
    df_cm_report = pd.concat([df_cm_tft, df_cm_lstm])

    def calculate_advanced_metrics(cm, actual_moves, pred_moves):
        # 1. 상승(악화) 클래스의 Recall 및 Precision
        # cm 구조: [0:하락, 1:유지, 2:상승]
        actual_up_total = cm[2, :].sum() # 실제 '상승'인 전체 개수
        pred_up_correct = cm[2, 2]       # '상승'을 '상승'으로 맞춘 개수
        
        recall_up = pred_up_correct / actual_up_total if actual_up_total > 0 else 0
        
        # 2. 치명적 오분류율 (역전분류율: 실제 하락인데 상승 예측 + 실제 상승인데 하락 예측)
        severe_errors = cm[0, 2] + cm[2, 0]
        severe_error_rate = severe_errors / cm.sum()
        
        # 3. 변동 케이스(유지 제외) 한정 정확도
        # 실제 데이터가 '상승' 또는 '하락'인 경우만 필터링
        up_down_mask = actual_moves.isin(['상승', '하락'])
        if up_down_mask.sum() > 0:
            updown_acc = accuracy_score(actual_moves[up_down_mask], pred_moves[up_down_mask])
        else:
            updown_acc = 0.0
            
        return recall_up, severe_error_rate, updown_acc

    tft_metrics = calculate_advanced_metrics(cm_tft, res_df_raw['Actual_Move'], res_df_raw['TFT_Move'])
    lstm_metrics = calculate_advanced_metrics(cm_lstm, res_df_raw['Actual_Move'], res_df_raw['LSTM_Move'])

    df_raw_pfmc = pd.DataFrame({
        'model': ['TFT', 'LSTM'],
        'rmse': [tft_raw_rmse, lstm_raw_rmse],
        'mae': [tft_raw_mae, lstm_raw_mae],
        'directional_accuracy': [tft_acc, lstm_acc],
        'recall_UP_class': [tft_metrics[0], lstm_metrics[0]],        # 상승 클래스 재현율
        'severe_error_rate': [tft_metrics[1], lstm_metrics[1]],      # 치명적 오분류율
        'updown_case_accuracy': [tft_metrics[2], lstm_metrics[2]]    # 변동상황 적중률
    })    

    df_pfmc = pd.DataFrame({
        'tft_rmse': [tft_rmse],
        'tft_mae': [tft_mae],
        'lstm_rmse': [lstm_rmse],
        'lstm_mae': [lstm_mae]
    })

    # TFT 변수중요도 추출
    print('########## 중요도(1) 일반공변량 ##########')
    avg_weights = np.concatenate(all_weights).mean(axis=(0, 1))
    features_name = scaler.feature_names_in_
    imp_df = pd.DataFrame({
        'Feature': features_name,
        'Importance': avg_weights
    }).sort_values(by='Importance', ascending=False)

    print('########## 중요도(2) 시계열에 따른 일반공변량 ##########')
    concatenated_weights = np.concatenate(all_weights, axis=0)
    avg_weights_by_time = concatenated_weights.mean(axis=0)

    time_steps = [f't-{seq_length - i}' for i in range(seq_length)]

    imp_time_df = pd.DataFrame(
        avg_weights_by_time,
        columns=features_name,
        index=time_steps
    )

    print('########## 중요도(3) 전체 중에서 중요 공변량 ##########')
    concatenated_attn = np.concatenate(all_attention_score, axis=0)
    # (전체 샘플 수, 시계열 길이, d_model)
    avg_attn_to_last = concatenated_attn[:, -1, :].mean(axis=0)
    # 시계열 마지막 시점에 대한 attention score 평균 계산
    avg_vsn_weights = concatenated_weights.mean(axis=0)
    # vsn 포함한 평균 중요도 계산
    global_importance = avg_vsn_weights * avg_attn_to_last[:, np.newaxis]
    # 시계열 중요도와 attention score를 곱하여 글로벌 중요도 계산

    global_imp_df = pd.DataFrame(
        global_importance,
        columns=features_name,
        index=time_steps
    )

    print('########## 중요도(4) 산업별 성능 집계 ##########')
    def analyze_by_industry(res_df_raw, target, outdir):
        print(f'########## 산업별({grp_cd}) 성능 분석 중 ##########')

        industry_stats = []

        for loop_grp_cd, group in res_df_raw.groupby(grp_cd):
            # 1. 수치적 에러 (RMSE, MAE)
            rmse = np.sqrt(mean_squared_error(group['Actual'], group['TFT_Prediction']))
            mae = mean_absolute_error(group['Actual'], group['TFT_Prediction'])

            # 2. 방향성 정확도 (Hit Rate)
            # 실제 변동과 예측 변동이 일치하는 비율
            hit_count = (group['Actual_Move'] == group['TFT_Move']).sum()
            hit_rate = hit_count / len(group)

            industry_stats.append({
                'grp_cd': loop_grp_cd,
                'Sample_Count': len(group),
                'RMSE': rmse,
                'MAE': mae,
                'Hit_Rate': hit_rate
            })

        df_ind_pfmc = pd.DataFrame(industry_stats).sort_values(by='Hit_Rate', ascending=False)
        # df_ind_pfmc.to_csv(outdir + f'{target}_performance_by_industry.csv', index=False)

        return df_ind_pfmc

    df_ind_pfmc = analyze_by_industry(res_df_raw, target, outdir)

    print('########## 중요도(5) 산업별 변수중요도 ##########')
    sample_avg_weights = concatenated_weights.mean(axis=1)

    df_sample_weights = pd.DataFrame(sample_avg_weights, columns=features_name)
    df_sample_weights[grp_cd] = res_df_raw[grp_cd].values # test_ds 기준이므로 행 개수 동일

    industry_imp_df = df_sample_weights.groupby(grp_cd).mean()

    print("===========================Visualization 중===========================")
    # 시각화 차트 저장
    fig0, axes0 = plt.subplots(figsize=(8,6))
    sns.scatterplot(data=res_df_raw, x='Actual', y='TFT_Prediction',
                    ax=axes0, color='blue', alpha=0.6, markers='x', label='TFT')
    sns.scatterplot(data=res_df_raw, x='Actual', y='LSTM_Prediction',
                    ax=axes0, color='grey', alpha=0.6, markers='x', label='LSTM')

    combined_raw = pd.concat([res_df_raw['Actual']
                             , res_df_raw['TFT_Prediction']
                             , res_df_raw['LSTM_Prediction']])
    min_raw, max_raw = combined_raw.min(), combined_raw.max()
    axes0.plot([min_raw, max_raw], [min_raw, max_raw], 'r--', lw=2, label='Identity')

    axes0.set_title(f"(RAW) Actual vs Prediction ({target})")
    axes0.legend()
    axes0.grid(True, linestyle='--', alpha=0.5)

    plt.tight_layout()
    save_path = outdir + f'{target}_IMG0.png'
    plt.savefig(save_path)

    print('==========시각화(1):TFT only,(2):raw Scatterplot==========')
    fig1, axes1 = plt.subplots(1, 2, figsize=(15, 6))
    ax1_1 = axes1[0]
    ax1_2 = axes1[1]
    sns.scatterplot(data=res_df_raw, x='Actual', y='TFT_Prediction',
                    ax=ax1_1, color='blue', alpha=0.6, markers='x', label='TFT')
    min_val = res_df_raw['Actual'].min()
    max_val = res_df_raw['Actual'].max()

    ax1_1.plot([0, max_val], [0, max_val], 'r--', lw=2, label='Identity')

    ax1_1.set_title(f"Actual vs Prediction ({target})")
    ax1_1.legend()
    ax1_1.grid(True, linestyle='--', alpha=0.5)

    sns.scatterplot(data=res_df_raw, x='Actual', y='LSTM_Prediction',
                    ax=ax1_2, color='grey', alpha=0.6, markers='x', label='LSTM')
    sns.scatterplot(data=res_df_raw, x='Actual', y='TFT_Prediction',
                    ax=ax1_2, color='blue', alpha=0.6, markers='x', label='TFT')

    combined_raw = pd.concat([res_df_raw['Actual']
                             , res_df_raw['TFT_Prediction']
                             , res_df_raw['LSTM_Prediction']])
    min_raw, max_raw = combined_raw.min(), combined_raw.max()
    ax1_2.plot([min_raw, max_raw], [min_raw, max_raw], 'r--', lw=2, label='Identity')

    ax1_2.set_title(f"(RAW) Actual vs Prediction ({target})")
    ax1_2.legend()
    ax1_2.grid(True, linestyle='--', alpha=0.5)

    plt.tight_layout()
    save_path = outdir + f'{target}_IMG1.png'
    plt.savefig(save_path)

    print('==========시각화(2) 변수중요도==========')
    print('==========시각화(2-1) 평균 변수중요도==========')
    fig2, axes2 = plt.subplots(2, 2, figsize=(21, 12))

    ax2_1 = axes2[0,0]
    ax2_2 = axes2[0,1]
    ax2_3 = axes2[1,0]
    ax2_4 = axes2[1,1]
    ax2_1.set_title("TFT Top 20 Feature Importance")
    sns.barplot(data=imp_df.head(20), x='Importance', y='Feature', ax=ax2_1, palette='Reds')
    print('==========시각화(2-2) 시계열 변수중요도==========')
    top_20_features = imp_time_df.mean(axis=0).sort_values(ascending=False).head(20).index
    imp_time_top20 = imp_time_df[top_20_features]
    sns.heatmap(imp_time_top20.T, annot=False, cmap='Blues', ax=ax2_2)
    ax2_2.set_title("Top 20 Feature Importance by Time Lag")
    ax2_2.set_xlabel("Time Step")
    ax2_2.set_ylabel("Features")

    print('==========시각화(2-3) Temporal Attention Score==========')
    ax2_3.plot(time_steps, avg_attn_to_last, marker='o', color='teal', linewidth=2, markersize=4)
    ax2_3.fill_between(time_steps, avg_attn_to_last, alpha=0.3, color='teal')

    ax2_3.set_title("Temporal Attention: How far back the model looks",
                    fontsize=14, fontweight='bold')
    ax2_3.set_xlabel("Time Lag (Past -> Present)")
    ax2_3.set_ylabel("Attention Weight")

    tick_spacing = 6
    ax2_3.set_xticks(time_steps[::tick_spacing])
    ax2_3.set_xticklabels(time_steps[::tick_spacing], rotation=45, ha='right')

    ax2_3.grid(True, axis='y', linestyle='--', alpha=0.5)

    print('==========시각화(2-4) 글로벌 중요도==========')
    top_20_global = global_imp_df.sum(axis=0).sort_values(ascending=False).head(20).index
    plot_global_df = global_imp_df[top_20_global].T # 변수가 y축, 시간이 x축

    sns.heatmap(plot_global_df, annot=False, cmap='Blues', ax=ax2_4)
    ax2_4.set_title("Global Importance (Attn x VSN): Top 20 Features")
    ax2_4.set_xlabel("Time Step (Lag)")
    ax2_4.set_ylabel("Features")

    plt.tight_layout()

    save_path = outdir + f'{target}_IMG2.png'
    plt.savefig(save_path)

    print('==========시각화(3) Hit-rate(상/하위 10개)==========')
    fig3, axes3 = plt.subplots(1,2, figsize=(15, 8))
    ax3_1 = axes3[0]
    top_10_ind = df_ind_pfmc.head(10)
    sns.barplot(data=top_10_ind, x='Hit_Rate', y='grp_cd', ax=ax3_1, palette='Greens_r')
    ax3_1.set_title("Top 10 Industries by Hit Rate")
    ax3_1.set_xlim(0, 1.1)
    # 수치 라벨 추가
    for i, v in enumerate(top_10_ind['Hit_Rate']):
        ax3_1.text(v + 0.01, i, f'{v:.1%}', color='black', va='center', fontweight='bold')
    ax3_2 = axes3[1]
    bot_10_ind = df_ind_pfmc.tail(10)
    sns.barplot(data=bot_10_ind, x='Hit_Rate', y='grp_cd', ax=ax3_2, palette='Greens_r')
    ax3_2.set_title("Bottom 10 Industries by Hit Rate")
    ax3_2.set_xlim(0, 1.1)
    # 수치 라벨 추가
    for i, v in enumerate(bot_10_ind['Hit_Rate']):
        ax3_2.text(v + 0.01, i, f'{v:.1%}', color='black', va='center', fontweight='bold')

    plt.tight_layout()
    save_path = outdir + f'{target}_IMG3.png'
    plt.savefig(save_path)

    print('==========시각화(4) 변동 방향성 혼동행렬==========')
    print('==========시각화(4-1) TFT==========')
    # fig4, axes4 = plt.subplots(1, 3, figsize=(21, 6))
    fig4, axes4 = plt.subplots(1, 2, figsize=(15, 6))
    ax4_1 = axes4[0]
    ax4_2 = axes4[1]
    # ax2_3 = axes4[2]

    # 성능 기록을 위해 라벨 순서 고정
    sns.heatmap(cm_tft/cm_tft.sum(), fmt='.2%', annot=True, cmap='Blues', ax=ax4_1,
                xticklabels=move_labels, yticklabels=move_labels)
    ax4_1.set_title(f"***TFT Dir-Accuracy: {acc_tft:.2%} (Thr: {int(threshold*100)}%) \n , 역전분류 : {(cm_tft[2,0]+cm_tft[0,2])/cm_tft.sum():.1%}***")
    ax4_1.set_ylabel("Actual Move")
    ax4_1.set_xlabel("Predicted Move(TFT)")

    print('==========시각화(4-2) LSTM==========')
    # 성능 기록을 위해 라벨 순서 고정
    sns.heatmap(cm_lstm/cm_lstm.sum(), fmt='.2%', annot=True, cmap='Blues', ax=ax4_2,
                xticklabels=move_labels, yticklabels=move_labels)
    ax4_2.set_title(f"***LSTM Dir-Accuracy: {acc_lstm:.2%} (Thr: {int(threshold*100)}%) \n , 역전분류 : {(cm_lstm[2,0]+cm_lstm[0,2])/cm_lstm.sum():.1%}***")
    ax4_2.set_ylabel("Actual Move")
    ax4_2.set_xlabel("Predicted Move(LSTM)")

    plt.tight_layout()
    save_path = outdir + f'{target}_IMG4.png'
    plt.savefig(save_path)

    print('==========시각화(5) 변수중요도_산업별==========')
    fig5, axes5 = plt.subplots(1, 1, figsize=(8, 6))
    ax5 = axes5

    sns.heatmap(industry_imp_df, annot=False, cmap='Blues', ax=ax5)
    ax5.set_title("Feature Importance by Industry Group")
    ax5.set_xlabel("Features")
    ax5.set_ylabel("Industry")

    plt.tight_layout()
    save_path = outdir + f'{target}_IMG5.png'
    plt.savefig(save_path)

    print('==========★★★★★최종output★★★★★==========')
    # Meta 정보
    # dir_source = \
    # r'E:\기업여신AX#03. 컨설팅_2영역\30. ML기반_전략_심사모델_고도화_및_승인전략_재구성\300. 산업특화ML\04. PYTHON\data/'

    # df_meta_grp = pd.read_csv(dir_source + 'meta_grp_csv.gz',sep='\t') # GRP 정보
    # df_meta_var = pd.read_csv(dir_source + 'meta_var_csv.gz',sep='\t') # 항목 정보

    # res_df_raw = res_df_raw.merge(df_meta_grp,how='left',on='grp_cd')
    # df_ind_pfmc = df_ind_pfmc.merge(df_meta_grp,how='left',on='grp_cd')
    # imp_df = imp_df.merge(df_meta_var,how='left',on = 'Feature')

    # 1. 데이터 시트 저장
    with pd.ExcelWriter(excel_path) as writer:
        res_df_raw.to_excel(writer, sheet_name='1.예측모형 결과', index=False)
        df_ind_pfmc.to_excel(writer, sheet_name='2.산업별 퍼포먼스', index=False)
        imp_df.to_excel(writer, sheet_name='3.변수중요도', index=False)
        imp_time_df.to_excel(writer, sheet_name='4.시계열 변수중요도', index=True)
        global_imp_df.to_excel(writer, sheet_name='5.글로벌 변수중요도', index=True)
        df_pfmc.to_excel(writer, sheet_name='6.성능지표', index=False)
        df_raw_pfmc.to_excel(writer, sheet_name='6.성능지표(RAW)')
        df_cm_report.to_excel(writer, sheet_name='6_1.혼동행렬(CM)', index=True)
        df_scaler.to_excel(writer, sheet_name='7.스케일러 정보', index=False)
        df_epoch.to_excel(writer, sheet_name='8.에포크 정보', index=False)
        df_config.to_excel(writer, sheet_name='9.모델 설정', index=False)
        industry_imp_df.to_excel(writer, sheet_name='10.산업별 주요변수')
        # df_meta_grp.to_excel(writer, sheet_name='meta_그룹정보')
        # df_meta_var.to_excel(writer, sheet_name='meta_변수정보')

    # 2. 시각화 시트
    wb = load_workbook(excel_path)
    if '0.시각화' not in wb.sheetnames:
        ws = wb.create_sheet('0.시각화', 0) # 맨 앞에 생성
    else:
        ws = wb['0.시각화']
    ws.append([f'Analysis Report for {target}'])
    img_files = [
        (f'{target}_IMG0.png', 'A2'),
        (f'{target}_IMG1.png', 'M2'),
        (f'{target}_IMG2.png', 'A31'),
        (f'{target}_IMG3.png', 'A87'),
        (f'{target}_IMG4.png', 'A124'),
        (f'{target}_IMG5.png', 'A152')
    ]

    for img_name, cell_pos in img_files:
        img_path = os.path.join(outdir, img_name)
        if os.path.exists(img_path):
            img = OpenpyxlImage(img_path)
            ws.add_image(img, cell_pos)
    wb.save(excel_path)
    plt.show()