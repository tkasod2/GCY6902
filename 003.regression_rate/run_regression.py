# -*- coding: utf-8 -*-
import torch
import os
import gc
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

import matplotlib.pyplot as plt
import seaborn as sns

from torch.utils.data import DataLoader

# 기존 파일에서 클래스 및 함수 불러오기
from tft_model import TFTConfig, TemporalFusionTransformer
from train_tft import SeqDataset, collate_fn, train, train_lstm
from sklearn.metrics import mean_squared_error, mean_absolute_error
from sklearn.metrics import accuracy_score, f1_score, roc_auc_score, confusion_matrix, roc_curve, auc
from sklearn.preprocessing import label_binarize

# sample test
# from make_sample_df import *
from prepare_data import *
# 모델날짜
today_ = datetime.now().strftime("%y%m%d")
def fit_and_out(df,
                target,
                outdir,
                config_dict: dict,
                var_list: list,
                grp_cd="Symbol",
                epochs=2,
                patience=1,
                threshold=0.05,  # 변동 방향성 판단 기준 (예: 5% 이상 상승/하락)
                trk_start_dt='2024-01-01',
                seq_length=60,  # 60일
                dt_cut='2025-12-31'):

    x_past, x_known, x_static, y, scaler, le = prepare_data(
        df,
        target=target,
        var_list=var_list,
        trk_start_dt=trk_start_dt,
        seq_length=seq_length,
        dt_cut=dt_cut,
    )

    df_scaler = pd.DataFrame({
        "VAR": scaler.feature_names_in_,
        "mean": scaler.mean_,
        "std": scaler.scale_
    })

    train_ds = SeqDataset(x_past[0], y[0], x_known[0], x_static[0])
    valid_ds = SeqDataset(x_past[1], y[1], x_known[1], x_static[1])
    test_ds = SeqDataset(x_past[2], y[2], x_known[2], x_static[2])
    
    trn_loader = DataLoader(train_ds, batch_size=16, shuffle=True, collate_fn=collate_fn)
    val_loader = DataLoader(valid_ds, batch_size=16, shuffle=False, collate_fn=collate_fn)
    ts_loader = DataLoader(test_ds, batch_size=16, shuffle=False, collate_fn=collate_fn)
    print('-----------------------------------------------------------')
    print(f"Input X_past mean: {x_past[2].mean()}") 

    # 2. 역변환에 사용되는 mean, std 비교
    df_scaler = pd.DataFrame({"VAR": scaler.feature_names_in_, "mean": scaler.mean_, "std": scaler.scale_})
    print(df_scaler[df_scaler['VAR'] == target])
    print('-----------------------------------------------------------')
    def fit_model(config_dict,epochs=2,patience=1):
        cfg = TFTConfig(
            **config_dict
        )

        print("======1. TFT 모델 학습중======")
        model, df_epoch_tft = train(
            cfg=cfg,
            train_loader=trn_loader,
            valid_loader=val_loader,
            test_loader=ts_loader,
            epochs=epochs,#
            lr=1e-3,
            patience=patience,
            save_path=outdir + f'{target}_tft_test_{today_}.pt'
        )

        print("======2. LSTM 모델 학습중======")
        model_lstm, df_epoch_lstm = train_lstm(
            input_dim=cfg.past_vars,
            train_loader=trn_loader,
            valid_loader=val_loader,
            test_loader=ts_loader,
            epochs=epochs,#
            patience=patience,
            save_path=outdir + f'{target}_lstm_test_{today_}.pt'
        )
        print(config_dict)


        df_epoch_tft['gb'] = 'TFT'
        df_epoch_lstm['gb'] = 'LSTM'
        df_epoch = pd.concat([df_epoch_tft, df_epoch_lstm], axis=0, ignore_index=True)
        df_config = pd.DataFrame.from_dict({'keys': config_dict.keys(),'values': config_dict.values()})
        
        return model, model_lstm, df_epoch, df_config
    model, model_lstm, df_epoch, df_config = fit_model(config_dict,epochs=epochs,patience=patience)


    print("====================output building 중====================")

    model.eval()
    model_lstm.eval()
    with torch.no_grad():

        all_actuals, all_tft_preds, all_lstm_preds, all_weights,all_attention_score = [], [], [], [], []

        for x_p, x_k, x_s, p, y in ts_loader:
            # tft
            y_tft, aux = model(x_p, x_known=x_k, x_static=x_s, pos=p)
            # lstm
            y_lstm, _ = model_lstm(x_p)

            all_actuals.append(y.numpy())
            all_tft_preds.append(y_tft.numpy())
            all_lstm_preds.append(y_lstm.numpy())
            all_weights.append(aux['w_past'].numpy())
            all_attention_score.append(aux['attn_w'].numpy())

    # 결과 data frame
    res_df = pd.DataFrame({
        'TIME_IDX': [x+1 for x in test_ds.x_known[:,-1,:]], # 예측시점 : 이전 time window 기간의 max +1 처리
        grp_cd: le.inverse_transform(test_ds.x_static.reshape(-1,)),
        'Actual': np.concatenate(all_actuals).flatten(),
        'TFT_Prediction': np.concatenate(all_tft_preds).flatten(),
        'LSTM_Prediction': np.concatenate(all_lstm_preds).flatten()
    })

    # res_df.to_csv(outdir + f"{target}_comparison_results.csv", index=False)
    print('==========RAW기준으로 데이터 복원==========')
    ynum = var_list.index(target)
    res_df_raw = res_df.copy()
    df_scaler = pd.DataFrame({
        "VAR": scaler.feature_names_in_,
        "mean": scaler.mean_,
        "std": scaler.scale_
    })
    mean_scale = df_scaler.loc[df_scaler['VAR'] == target, 'mean'].item()
    std_scale = df_scaler.loc[df_scaler['VAR'] == target, 'std'].item()
    res_df_raw['Actual'] = res_df_raw['Actual'] * std_scale + mean_scale
    res_df_raw['TFT_Prediction'] = res_df_raw['TFT_Prediction'] * std_scale + mean_scale
    res_df_raw['LSTM_Prediction'] = res_df_raw['LSTM_Prediction'] * std_scale + mean_scale
    y_t_minus_1_raw = x_past[2][:, -1, ynum] # 테스트셋의 t-1 시점 원본값
    y_t_minus_1_raw = y_t_minus_1_raw * std_scale + mean_scale

    res_df_raw['Actual_t_minus_1'] = y_t_minus_1_raw
    threshold = threshold # +-5% 기준
    def classify_movement(current, base, thr):
        diff = (current - base) / (base + 1e-9) # 분모 0 방지
        if diff > thr: return "상승"
        elif diff < -thr: return "하락"
        else: return "유지"

    # 실제 변동(Actual t vs Actual t-1)
    res_df_raw['Actual_Move'] = res_df_raw.apply(
        lambda x: classify_movement(x['Actual'], x['Actual_t_minus_1'], threshold), axis=1)

    # TFT 예측 변동(TFT Pred t vs Actual t-1)
    res_df_raw['TFT_Move'] = res_df_raw.apply(
        lambda x: classify_movement(x['TFT_Prediction'], x['Actual_t_minus_1'], threshold), axis=1)

    # LSTM 예측 변동(LSTM Pred t vs Actual t-1)
    res_df_raw['LSTM_Move'] = res_df_raw.apply(
        lambda x: classify_movement(x['LSTM_Prediction'], x['Actual_t_minus_1'], threshold), axis=1)
    
    # Confusion Matrix 계산
    move_labels = ["하락", "유지", "상승"]
    cm_tft = confusion_matrix(res_df_raw['Actual_Move'], res_df_raw['TFT_Move'], labels=move_labels)
    cm_lstm = confusion_matrix(res_df_raw['Actual_Move'], res_df_raw['LSTM_Move'], labels=move_labels)

    # 정확도 계산
    acc_tft = accuracy_score(res_df_raw['Actual_Move'], res_df_raw['TFT_Move'])
    acc_lstm = accuracy_score(res_df_raw['Actual_Move'], res_df_raw['LSTM_Move'])

    print(f"TFT Directional Accuracy: {acc_tft:.4f}")
    print(f"LSTM Directional Accuracy: {acc_lstm:.4f}")

    actual = res_df['Actual']
    tft_pred = res_df['TFT_Prediction']
    lstm_pred = res_df['LSTM_Prediction']

    tft_rmse = np.sqrt(mean_squared_error(actual, tft_pred))
    tft_mae = mean_absolute_error(actual, tft_pred)

    lstm_rmse = np.sqrt(mean_squared_error(actual, lstm_pred))
    lstm_mae = mean_absolute_error(actual, lstm_pred)

    print(tft_rmse, tft_mae, lstm_rmse, lstm_mae)

    df_pfmc = pd.DataFrame({
        'tft_rmse': [tft_rmse],
        'tft_mae': [tft_mae],
        'lstm_rmse': [lstm_rmse],
        'lstm_mae': [lstm_mae]
    })

    # TFT 변수중요도 추출
    print('########## 중요도(1) 일반공변량 ##########')
    avg_weights = np.concatenate(all_weights).mean(axis=(0, 1))
    features_name = scaler.feature_names_in_
    imp_df = pd.DataFrame({
        'Feature': features_name,
        'Importance': avg_weights
    }).sort_values(by='Importance', ascending=False)

    print('########## 중요도(2) 시계열에 따른 일반공변량 ##########')
    concatenated_weights = np.concatenate(all_weights, axis=0)
    avg_weights_by_time = concatenated_weights.mean(axis=0)

    time_steps = [f't-{seq_length - i}' for i in range(seq_length)]

    imp_time_df = pd.DataFrame(
        avg_weights_by_time,
        columns=features_name,
        index=time_steps
    )
    imp_time_df.to_csv(outdir + f"{target}_feature_importance_by_time.csv")

    print('########## 중요도(3) 전체 중에서 중요 공변량 ##########')
    concatenated_attn = np.concatenate(all_attention_score, axis=0) # (전체 샘플 수, 시계열 길이, d_model)
    avg_attn_to_last = concatenated_attn[:, -1, :].mean(axis=0) # 시계열 마지막 시점에 대한 attention score 평균 계산
    avg_vsn_weights = concatenated_weights.mean(axis=0) # vsn 포함한 평균 중요도 계산
    global_importance = avg_vsn_weights * avg_attn_to_last[:, np.newaxis] # 시계열 중요도와 attention score를 곱하여 글로벌 중요도 계산

    global_imp_df = pd.DataFrame(
        global_importance,
        columns=features_name,
        index=time_steps
    )

    print('########## 중요도(4) 산업별 성능 집계 ##########')
    def analyze_by_industry(res_df_raw, target, outdir):
        print('########## 산업별(grp_cd) 성능 분석 중 ##########')
        
        industry_stats = []
        
        for loop_grp_cd, group in res_df_raw.groupby(grp_cd):
            # 1. 수치적 에러 (RMSE, MAE)
            rmse = np.sqrt(mean_squared_error(group['Actual'], group['TFT_Prediction']))
            mae = mean_absolute_error(group['Actual'], group['TFT_Prediction'])
            
            # 2. 방향성 정확도 (Hit Rate)
            # 실제 변동과 예측 변동이 일치하는 비율
            hit_count = (group['Actual_Move'] == group['TFT_Move']).sum()
            hit_rate = hit_count / len(group)
            
            industry_stats.append({
                grp_cd: loop_grp_cd,
                'Sample_Count': len(group),
                'RMSE': rmse,
                'MAE': mae,
                'Hit_Rate': hit_rate
            })
        
        df_ind_pfmc = pd.DataFrame(industry_stats).sort_values(by='Hit_Rate', ascending=False)
        df_ind_pfmc.to_csv(outdir + f"{target}_performance_by_industry.csv", index=False)
        
        return df_ind_pfmc

    df_ind_pfmc = analyze_by_industry(res_df_raw, target, outdir)

    print("============================== Visualization中 ==============================")
    # 시각화 차트 저장
    fig1, axes1 = plt.subplots(1, 2, figsize=(15, 6))
    ax1 = axes1[0]
    ax2 = axes1[1]

    print('==========시각화(1) Scatter Plot==========')
    sns.scatterplot(data=res_df, x='Actual', y='TFT_Prediction',
                    ax=ax1, color='blue', alpha=0.6, markers='x', label='TFT')
    sns.scatterplot(data=res_df, x='Actual', y='LSTM_Prediction',
                    ax=ax1, color='grey', alpha=0.6, markers='x', label='LSTM')
    min_val = res_df['Actual'].min()
    max_val = res_df['Actual'].max()

    ax1.plot([0, max_val], [0, max_val], 'r--', lw=2, label='Identity')

    ax1.set_title(f"Actual vs Prediction ({target})")
    ax1.legend()
    ax1.grid(True, linestyle='--', alpha=0.5)

    print('==========시각화(2) Scatter Plot(RAW)==========')
    sns.scatterplot(data=res_df_raw, x='Actual', y='TFT_Prediction',
                    ax=ax2, color='blue', alpha=0.6, markers='x', label='TFT')
    sns.scatterplot(data=res_df_raw, x='Actual', y='LSTM_Prediction',
                    ax=ax2, color='grey', alpha=0.6, markers='x', label='LSTM')
    
    combined_raw = pd.concat([res_df_raw['Actual'], res_df_raw['TFT_Prediction'], res_df_raw['LSTM_Prediction']])
    min_raw, max_raw = combined_raw.min(), combined_raw.max()
    ax2.plot([min_raw, max_raw], [min_raw, max_raw], 'r--', lw=2, label='Identity')

    ax2.set_title(f"(RAW) Actual vs Prediction ({target})")
    ax2.legend()
    ax2.grid(True, linestyle='--', alpha=0.5)
    
    plt.tight_layout()
    save_path = outdir + f"{target}_IMG1.png"
    plt.savefig(save_path)

    print('==========시각화(3) 시계열 변수중요도==========')
    fig2, axes2 = plt.subplots(1, 2, figsize=(15, 6))
    ax3 = axes2[0]
    ax4 = axes2[1]

    top_20_features = imp_time_df.mean(axis=0).sort_values(ascending=False).head(20).index
    imp_time_top20 = imp_time_df[top_20_features]
    sns.heatmap(imp_time_top20.T, annot=False, cmap='YlGnBu', ax=ax3)
    ax3.set_title("Top 20 Feature Importance by Time Lag")
    ax3.set_xlabel("Time Step")
    ax3.set_ylabel("Features")

    print('==========시각화(4) 글로벌 중요도==========')
    top_20_global = global_imp_df.sum(axis=0).sort_values(ascending=False).head(20).index
    plot_global_df = global_imp_df[top_20_global].T # 변수가 y축, 시간이 x축

    sns.heatmap(plot_global_df, annot=False, cmap='YlGnBu', ax=ax4)
    ax4.set_title("Global Importance (Attn x VSN): Top 20 Features")
    ax4.set_xlabel("Time Step (Lag)")
    ax4.set_ylabel("Features")

    plt.tight_layout()
    save_path = outdir + f"{target}_IMG2.png"
    plt.savefig(save_path)    

    print('==========시각화(5) 평균 변수중요도==========')
    fig3, axes3 = plt.subplots(1, 2, figsize=(15, 6))
    ax5 = axes3[0]
    ax6 = axes3[1]
    sns.barplot(data=imp_df.head(20), x='Importance', y='Feature', ax=ax5, palette='Reds')
    ax5.set_title("TFT Top 20 Feature Importance")

    print('==========시각화(6) 산업별 중요도(10개 sample)==========')
    top_10_ind = df_ind_pfmc.head(10)
    sns.barplot(data=top_10_ind, x='Hit_Rate', y=grp_cd, ax=ax6, palette='Greens_r')
    ax6.set_title("Top 10 Industries by Hit Rate")
    ax6.set_xlim(0, 1.1)
    # 수치 라벨 추가
    for i, v in enumerate(top_10_ind['Hit_Rate']):
        ax6.text(v + 0.01, i, f'{v:.1%}', color='black', va='center', fontweight='bold')
    plt.tight_layout()
    save_path = outdir + f"{target}_IMG3.png"
    plt.savefig(save_path)    

    print('==========시각화(7) 변동 방향성 혼동행렬(TFT)==========')
    fig4, axes4 = plt.subplots(1, 3, figsize=(21, 6))
    ax7 = axes4[0]
    ax8 = axes4[1]
    ax9 = axes4[2]

    # 성능 기록을 위해 라벨 순서 고정
    sns.heatmap(cm_tft / cm_tft.sum(axis=1)[:, None], annot=True, fmt=".1%", cmap='Blues', ax=ax7,
            xticklabels=move_labels, yticklabels=move_labels)
    ax7.set_title(f"TFT Dir-Accuracy: {acc_tft:.2%} (Thr: {int(threshold*100)}%)")
    ax7.set_ylabel("Actual Move")
    ax7.set_xlabel("Predicted Move(TFT)")

    print('==========시각화(8) 변동 방향성 혼동행렬(LSTM)==========')
    # 성능 기록을 위해 라벨 순서 고정
    sns.heatmap(cm_lstm / cm_lstm.sum(axis=1)[:, None], annot=True, fmt=".1%", cmap='Blues', ax=ax8,
            xticklabels=move_labels, yticklabels=move_labels)
    ax8.set_title(f"LSTM Dir-Accuracy: {acc_lstm:.2%} (Thr: {int(threshold*100)}%)")
    ax8.set_ylabel("Actual Move")
    ax8.set_xlabel("Predicted Move(LSTM)")

    print('==========시각화(9) Temporal Attention Score==========')
    ax9.plot(time_steps, avg_attn_to_last, marker='o', color='teal', linewidth=2, markersize=4)
    ax9.fill_between(time_steps, avg_attn_to_last, alpha=0.3, color='teal')
    
    ax9.set_title("Temporal Attention: How far back the model looks", fontsize=14, fontweight='bold')
    ax9.set_xlabel("Time Lag (Past -> Present)")
    ax9.set_ylabel("Attention Weight")

    tick_spacing = 6
    ax9.set_xticks(time_steps[::tick_spacing])
    ax9.set_xticklabels(time_steps[::tick_spacing], rotation=45, ha='right')
    
    ax9.grid(True, axis='y', linestyle='--', alpha=0.5)

    plt.tight_layout()
    save_path = outdir + f"{target}_IMG4.png"
    plt.savefig(save_path)    

    excel_path = outdir + f"{target}_final_report_{today_}.xlsx"
    
    with pd.ExcelWriter(excel_path) as writer:
        # 1. 데이터 시트 저장
        res_df_raw.to_excel(writer, sheet_name='1.예측모형 결과', index=False)
        df_ind_pfmc.to_excel(writer, sheet_name='2.산업별 퍼포먼스', index=False)
        imp_df.to_excel(writer, sheet_name='3.변수중요도', index=False)
        imp_time_df.to_excel(writer, sheet_name='4.시계열 변수중요도', index=True)
        global_imp_df.to_excel(writer, sheet_name='5.글로벌 변수중요도', index=True)
        df_pfmc.to_excel(writer, sheet_name='6.성능지표', index=False)
        df_scaler.to_excel(writer, sheet_name='7.스케일러 정보', index=False)
        df_epoch.to_excel(writer, sheet_name='8.에포크 정보', index=False)
        df_config.to_excel(writer, sheet_name='9.모델 설정', index=False)
        
    # 2. 시각화 시트
    
    wb = load_workbook(excel_path)
    ws = wb.create_sheet('0.시각화', 0)
    img_files = [
        (f"{target}_IMG1.png", 'A3'),
        (f"{target}_IMG2.png", 'A25'),
        (f"{target}_IMG3.png", 'A47'),
        (f"{target}_IMG4.png", 'A69')
    ]
    for img_name, cell_pos in img_files:
        img_path = os.path.join(outdir, img_name)
        if os.path.exists(img_path):
            img = OpenpyxlImage(img_path)
            # 이미지 크기 조절 (선택 사항)
            # img.width = img.width * 0.7
            # img.height = img.height * 0.7
            ws.add_image(img, cell_pos)
    wb.save(excel_path)

    plt.show()
